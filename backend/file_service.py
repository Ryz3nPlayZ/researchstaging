"""
File Service

Handles file and folder operations for project workspaces.
Supports upload, download, CRUD operations, and organization.
Uses storage abstraction layer for backend-agnostic file operations.
"""
import os
import hashlib
import logging
import mmap
import re
from typing import Optional, List, Dict, Any
from pathlib import Path
from fastapi import UploadFile, HTTPException
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from database import Folder, File, Project
from database.models import utc_now
from storage_service import get_storage

logger = logging.getLogger(__name__)

# Configuration
UPLOAD_DIR = os.environ.get("UPLOAD_DIR", "uploads")
PROJECTS_DIR = os.path.join(UPLOAD_DIR, "projects")
MAX_FILE_SIZE = int(os.environ.get("MAX_FILE_SIZE", str(50 * 1024 * 1024)))  # 50MB default

# Ensure upload directory exists (for local storage fallback)
os.makedirs(PROJECTS_DIR, exist_ok=True)

# Supported file types and their mime types
SUPPORTED_FILE_TYPES = {
    # Documents
    ".pdf": ["application/pdf"],
    ".docx": ["application/vnd.openxmlformats-officedocument.wordprocessingml.document"],
    ".md": ["text/markdown", "text/plain"],

    # Code files
    ".py": ["text/x-python", "text/plain"],
    ".r": ["text/x-r", "text/plain"],
    ".js": ["text/javascript", "application/javascript", "text/plain"],

    # Data files
    ".csv": ["text/csv", "application/csv", "text/plain"],
    ".xlsx": ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
              "application/vnd.ms-excel",
              "application/octet-stream"],
}

# Extensions for validation
SUPPORTED_EXTENSIONS = set(SUPPORTED_FILE_TYPES.keys())


# ============== Custom Exceptions ==============

class FileServiceError(HTTPException):
    """Base exception for file service errors."""
    def __init__(self, status_code: int, detail: str):
        super().__init__(status_code=status_code, detail=detail)
        logger.error(f"FileServiceError: {detail}")


class UnsupportedFileTypeError(FileServiceError):
    """Raised when file type is not supported."""
    def __init__(self, filename: str, file_type: str):
        super().__init__(
            status_code=400,
            detail=f"Unsupported file type: {file_type}. Supported types: {', '.join(SUPPORTED_EXTENSIONS)}"
        )
        logger.warning(f"Unsupported file type: {filename} ({file_type})")


class FileTooLargeError(FileServiceError):
    """Raised when file exceeds size limit."""
    def __init__(self, filename: str, size: int, max_size: int):
        size_mb = size / (1024 * 1024)
        max_mb = max_size / (1024 * 1024)
        super().__init__(
            status_code=413,
            detail=f"File '{filename}' ({size_mb:.2f}MB) exceeds maximum size ({max_mb:.2f}MB)"
        )
        logger.warning(f"File too large: {filename} ({size} bytes)")


class FileMetadataExtractionError(FileServiceError):
    """Raised when metadata extraction fails."""
    def __init__(self, filename: str, reason: str):
        super().__init__(
            status_code=500,
            detail=f"Failed to extract metadata from '{filename}': {reason}"
        )
        logger.error(f"Metadata extraction failed: {filename} - {reason}")


# ============== Path Utilities ==============

def _get_project_storage_path(project_id: str) -> str:
    """Get the storage path for a project."""
    return os.path.join(PROJECTS_DIR, project_id)


def _get_file_storage_key(project_id: str, file_id: str, extension: str) -> str:
    """
    Get the storage key for a specific file.
    This key is used by the storage backend (local, S3, or R2).
    Format: projects/{project_id}/{subdir}/{file_id}.{extension}
    """
    subdir = file_id[:2]
    return f"projects/{project_id}/{subdir}/{file_id}.{extension}"


def _get_file_storage_path(project_id: str, file_id: str, extension: str) -> str:
    """Get the local storage path for a specific file (legacy, for compatibility)."""
    project_path = _get_project_storage_path(project_id)
    subdir = file_id[:2]
    return os.path.join(project_path, subdir, f"{file_id}.{extension}")


# ============== File Type Validation ==============

def _get_file_extension(filename: str) -> str:
    """Extract file extension from filename."""
    _, ext = os.path.splitext(filename.lower())
    return ext


def _is_supported_file_type(extension: str) -> bool:
    """Check if file extension is supported."""
    return extension in SUPPORTED_EXTENSIONS


def _normalize_mime_type(mime_type: Optional[str]) -> str:
    """Normalize mime type for consistency."""
    if not mime_type:
        return "application/octet-stream"
    return mime_type.lower().strip()


def _validate_file_type(filename: str, declared_mime: Optional[str] = None) -> tuple[str, List[str]]:
    """
    Validate file type and return extension with allowed mime types.

    Args:
        filename: Name of the file
        declared_mime: Declared mime type from upload

    Returns:
        Tuple of (extension, allowed_mime_types)

    Raises:
        UnsupportedFileTypeError: If file type is not supported
    """
    extension = _get_file_extension(filename)

    if not _is_supported_file_type(extension):
        raise UnsupportedFileTypeError(filename, extension)

    allowed_mimes = SUPPORTED_FILE_TYPES[extension]
    return extension, allowed_mimes


def _validate_file_size(content: bytes, filename: str) -> None:
    """
    Validate file size against maximum limit.

    Args:
        content: File content
        filename: Name of the file

    Raises:
        FileTooLargeError: If file exceeds size limit
    """
    size = len(content)
    if size > MAX_FILE_SIZE:
        raise FileTooLargeError(filename, size, MAX_FILE_SIZE)


# ============== Duplicate Handling ==============

async def _generate_unique_filename(
    db: AsyncSession,
    project_id: str,
    folder_id: Optional[str],
    filename: str,
    folder_path: str = ""
) -> tuple[str, str]:
    """
    Generate a unique filename by appending counter if duplicate exists.

    Args:
        db: Database session
        project_id: Project ID
        folder_id: Folder ID (if any)
        filename: Original filename
        folder_path: Path to folder

    Returns:
        Tuple of (display_name, storage_name)
    """
    name, ext = os.path.splitext(filename)
    counter = 0
    display_name = filename
    storage_name = filename

    # Check for duplicates
    while True:
        # Build path for checking
        test_path = f"{folder_path}/{storage_name}".replace("//", "/").strip("/")

        # Check if file with this path exists
        existing = await db.execute(
            select(File).where(
                File.project_id == project_id,
                File.path == test_path
            )
        )
        exists = existing.first()

        if not exists:
            # No duplicate found
            break

        # Duplicate found, increment counter
        counter += 1
        storage_name = f"{name} ({counter}){ext}"
        display_name = storage_name

    return display_name, storage_name


# ============== Metadata Extraction ==============

def _extract_pdf_metadata(filepath: str) -> Dict[str, Any]:
    """Extract metadata from PDF file."""
    try:
        import fitz  # PyMuPDF
        doc = fitz.open(filepath)
        metadata = {
            "page_count": doc.page_count,
            "title": doc.metadata.get("title", ""),
            "author": doc.metadata.get("author", ""),
            "subject": doc.metadata.get("subject", ""),
            "creator": doc.metadata.get("creator", ""),
        }
        doc.close()
        return metadata
    except Exception as e:
        logger.warning(f"PDF metadata extraction failed: {e}")
        return {"page_count": 0}


def _extract_csv_metadata(filepath: str) -> Dict[str, Any]:
    """Extract metadata from CSV file."""
    try:
        import pandas as pd
        df = pd.read_csv(filepath, nrows=0)
        # Count rows efficiently
        with open(filepath, 'rb') as f:
            # Use mmap for large files
            mm = mmap.mmap(f.fileno(), 0, access=mmap.ACCESS_READ)
            row_count = sum(1 for line in iter(mm.readline, b'')) - 1  # Subtract header
            mm.close()

        return {
            "row_count": max(0, row_count),
            "columns": list(df.columns),
            "column_count": len(df.columns)
        }
    except Exception as e:
        logger.warning(f"CSV metadata extraction failed: {e}")
        return {"row_count": 0, "column_count": 0, "columns": []}


def _extract_excel_metadata(filepath: str) -> Dict[str, Any]:
    """Extract metadata from Excel file."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

        sheets_info = {}
        total_rows = 0

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            row_count = sheet.max_row
            col_count = sheet.max_column
            sheets_info[sheet_name] = {
                "row_count": row_count,
                "column_count": col_count
            }
            total_rows = max(total_rows, row_count)

        wb.close()
        return {
            "sheet_names": wb.sheetnames,
            "sheet_count": len(wb.sheetnames),
            "total_rows": total_rows,
            "sheets": sheets_info
        }
    except Exception as e:
        logger.warning(f"Excel metadata extraction failed: {e}")
        return {"sheet_count": 0, "sheet_names": [], "sheets": {}}


def _extract_code_metadata(filepath: str, extension: str) -> Dict[str, Any]:
    """Extract metadata from code files."""
    try:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            line_count = sum(1 for _ in f)

        language_map = {
            ".py": "python",
            ".r": "r",
            ".js": "javascript"
        }

        return {
            "line_count": line_count,
            "language": language_map.get(extension, "unknown")
        }
    except Exception as e:
        logger.warning(f"Code metadata extraction failed: {e}")
        return {"line_count": 0, "language": "unknown"}


def _extract_markdown_metadata(filepath: str) -> Dict[str, Any]:
    """Extract metadata from Markdown file."""
    try:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
            line_count = len(content.splitlines())
            word_count = len(content.split())
            char_count = len(content)

            # Extract front matter if present
            title = None
            frontmatter_match = re.match(r'^---\n(.*?)\n---', content, re.DOTALL)
            if frontmatter_match:
                frontmatter = frontmatter_match.group(1)
                title_match = re.search(r'title:\s*(.+)', frontmatter)
                if title_match:
                    title = title_match.group(1).strip('"\'')

            return {
                "line_count": line_count,
                "word_count": word_count,
                "char_count": char_count,
                "title": title
            }
    except Exception as e:
        logger.warning(f"Markdown metadata extraction failed: {e}")
        return {"line_count": 0, "word_count": 0, "char_count": 0}


def _extract_file_metadata(filepath: str, extension: str) -> Dict[str, Any]:
    """
    Extract metadata from file based on type.

    Args:
        filepath: Path to file
        extension: File extension

    Returns:
        Dictionary with extracted metadata
    """
    metadata = {}

    try:
        if extension == ".pdf":
            metadata = _extract_pdf_metadata(filepath)
        elif extension == ".csv":
            metadata = _extract_csv_metadata(filepath)
        elif extension == ".xlsx":
            metadata = _extract_excel_metadata(filepath)
        elif extension in [".py", ".r", ".js"]:
            metadata = _extract_code_metadata(filepath, extension)
        elif extension == ".md":
            metadata = _extract_markdown_metadata(filepath)
        else:
            # For other types, just return basic info
            metadata = {}

    except Exception as e:
        logger.error(f"Unexpected error extracting metadata: {e}")
        metadata = {}

    return metadata


# ============== CRUD Operations ==============

async def create_folder(
    db: AsyncSession,
    project_id: str,
    name: str,
    parent_folder_id: Optional[str] = None,
    description: Optional[str] = None
) -> Folder:
    """Create a new folder in a project."""
    # Verify project exists
    result = await db.execute(select(Project).where(Project.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise FileServiceError(status_code=404, detail="Project not found")

    # Generate unique path
    parent_path = ""
    if parent_folder_id:
        parent = await db.get(Folder, parent_folder_id)
        if not parent or parent.project_id != project_id:
            raise FileServiceError(status_code=400, detail="Invalid parent folder")
        parent_path = parent.path

    path = f"{parent_path}/{name}".replace("//", "/").strip("/")

    # Check for duplicate path
    existing = await db.execute(
        select(Folder).where(Folder.path == path)
    )
    if existing.first():
        raise FileServiceError(status_code=400, detail=f"Folder '{path}' already exists")

    folder = Folder(
        project_id=project_id,
        parent_folder_id=parent_folder_id,
        name=name,
        path=path,
        description=description
    )
    db.add(folder)
    await db.commit()
    await db.refresh(folder)

    logger.info(f"Created folder: {path}")
    return folder


async def upload_file(
    db: AsyncSession,
    project_id: str,
    file: UploadFile,
    folder_id: Optional[str] = None,
    description: Optional[str] = None
) -> File:
    """
    Upload a file to a project with validation, duplicate handling, and metadata extraction.

    Args:
        db: Database session
        project_id: Project ID
        file: UploadFile object
        folder_id: Optional folder ID
        description: Optional file description

    Returns:
        Created File record

    Raises:
        FileServiceError: On various error conditions
    """
    try:
        # Verify project exists
        result = await db.execute(select(Project).where(Project.id == project_id))
        project = result.scalar_one_or_none()
        if not project:
            raise FileServiceError(status_code=404, detail="Project not found")

        # Determine folder path
        folder_path = ""
        if folder_id:
            folder = await db.get(Folder, folder_id)
            if not folder or folder.project_id != project_id:
                raise FileServiceError(status_code=400, detail="Invalid folder")
            folder_path = folder.path

        # Get filename and extension
        filename = file.filename
        if not filename:
            raise FileServiceError(status_code=400, detail="Filename is required")

        # Validate file type
        try:
            extension, allowed_mimes = _validate_file_type(filename, file.content_type)
        except UnsupportedFileTypeError:
            raise
        except Exception as e:
            logger.error(f"File type validation error: {e}")
            raise FileServiceError(status_code=400, detail=f"Invalid filename: {filename}")

        # Read file content
        content = await file.read()

        # Validate file size
        try:
            _validate_file_size(content, filename)
        except FileTooLargeError:
            raise

        # Reset file pointer for potential re-reading
        await file.seek(0)

        # Calculate content hash
        content_hash = hashlib.sha256(content).hexdigest()

        # Generate file ID
        from database import generate_uuid
        file_id = generate_uuid()

        # Handle duplicate filenames
        display_name, storage_name = await _generate_unique_filename(
            db, project_id, folder_id, filename, folder_path
        )

        # Generate path
        path = f"{folder_path}/{storage_name}".replace("//", "/").strip("/")

        # Get storage backend
        storage = get_storage()

        # Store file using storage backend
        storage_key = _get_file_storage_key(project_id, file_id, extension)
        await storage.upload_file(storage_key, content, metadata={"filename": filename})

        # For local storage, also maintain legacy storage_path for compatibility
        storage_path = _get_file_storage_path(project_id, file_id, extension)

        # Extract metadata (save to temp file for metadata extraction)
        try:
            # For metadata extraction, save to temp file
            import tempfile
            with tempfile.NamedTemporaryFile(delete=False, suffix=extension) as tmp_file:
                tmp_file.write(content)
                tmp_path = tmp_file.name

            metadata = _extract_file_metadata(tmp_path, extension)

            # Clean up temp file
            try:
                os.unlink(tmp_path)
            except Exception as cleanup_error:
                logger.warning(f"Failed to clean up temp file {tmp_path}: {cleanup_error}")

        except Exception as e:
            logger.warning(f"Metadata extraction failed for {filename}: {e}")
            metadata = {}

        # Normalize mime type
        mime_type = _normalize_mime_type(file.content_type)

        # Create file record
        file_record = File(
            id=file_id,
            project_id=project_id,
            folder_id=folder_id,
            name=display_name,
            file_type=extension.lstrip("."),
            mime_type=mime_type,
            size_bytes=len(content),
            path=path,
            description=description,
            storage_path=storage_path,
            content_hash=content_hash,
            tags=metadata  # Store metadata in tags field for now
        )
        db.add(file_record)
        await db.commit()
        await db.refresh(file_record)

        logger.info(
            f"Uploaded file: {path} ({len(content)} bytes, {extension}, "
            f"metadata: {list(metadata.keys())})"
        )
        return file_record

    except (UnsupportedFileTypeError, FileTooLargeError, FileServiceError):
        # Re-raise known exceptions
        raise
    except Exception as e:
        logger.error(f"Unexpected error uploading file: {e}", exc_info=True)
        # Rollback transaction
        await db.rollback()
        raise FileServiceError(status_code=500, detail=f"Failed to upload file: {str(e)}")


async def list_project_files(
    db: AsyncSession,
    project_id: str
) -> List[dict]:
    """List all files and folders in a project."""
    # Get folders
    result = await db.execute(
        select(Folder).where(Folder.project_id == project_id)
    )
    folders = result.scalars().all()

    # Get files
    result = await db.execute(
        select(File).where(File.project_id == project_id)
    )
    files = result.scalars().all()

    # Convert to dict format
    folder_list = []
    file_list = []

    for folder in folders:
        folder_list.append({
            "id": folder.id,
            "name": folder.name,
            "type": "folder",
            "path": folder.path,
            "description": folder.description,
            "file_count": folder.file_count,
            "created_at": folder.created_at.isoformat()
        })

    for file in files:
        file_dict = {
            "id": file.id,
            "name": file.name,
            "type": "file",
            "fileType": file.file_type,
            "path": file.path,
            "description": file.description,
            "size_bytes": file.size_bytes,
            "created_at": file.created_at.isoformat(),
            "mime_type": file.mime_type
        }
        # Add metadata if present
        if file.tags:
            file_dict["metadata"] = file.tags
        file_list.append(file_dict)

    # Organize into tree structure
    def build_tree(parent_path=""):
        tree = []
        items = [f for f in folder_list if f["path"].startswith(parent_path)]
        for item in items:
            item["children"] = build_tree(item["path"])
            tree.append(item)
        # Add files in this directory
        dir_files = [f for f in file_list if f["path"].startswith(parent_path)]
        tree.extend(dir_files)
        return tree

    # Return flat list for now (can enhance to tree later)
    return folder_list + file_list


async def get_file(
    db: AsyncSession,
    file_id: str
) -> File:
    """Get a file by ID."""
    file = await db.get(File, file_id)
    if not file:
        raise FileServiceError(status_code=404, detail="File not found")
    return file


async def delete_file(
    db: AsyncSession,
    file_id: str
) -> None:
    """Delete a file."""
    file = await db.get(File, file_id)
    if not file:
        raise FileServiceError(status_code=404, detail="File not found")

    # Delete from storage backend
    storage = get_storage()
    try:
        # Extract storage key from storage_path
        # storage_path format: uploads/projects/{project_id}/{subdir}/{file_id}.{ext}
        # storage_key format: projects/{project_id}/{subdir}/{file_id}.{ext}
        if file.storage_path:
            # Convert local path to storage key
            storage_key = file.storage_path.replace(UPLOAD_DIR + "/", "").replace(UPLOAD_DIR, "")
            await storage.delete_file(storage_key)
    except Exception as e:
        logger.error(f"Failed to delete file from storage: {e}")

    # Also try to delete from local disk if it exists (for migration compatibility)
    if os.path.exists(file.storage_path):
        try:
            os.remove(file.storage_path)
        except Exception as e:
            logger.debug(f"Local file not found or already deleted: {e}")

    # Delete from database
    await db.delete(file)
    await db.commit()

    logger.info(f"Deleted file: {file.path}")


async def get_project_file_tree(
    db: AsyncSession,
    project_id: str
) -> dict:
    """Get the complete file tree for a project (for FileExplorer)."""
    # Get root folders (no parent)
    result = await db.execute(
        select(Folder).where(
            Folder.project_id == project_id,
            Folder.parent_folder_id.is_(None)
        ).order_by(Folder.name)
    )
    root_folders = result.scalars().all()

    # Build tree structure
    async def build_folder_node(folder_id: str) -> dict:
        folder = await db.get(Folder, folder_id)
        if not folder:
            return None

        # Get subfolders
        subfolder_result = await db.execute(
            select(Folder).where(
                Folder.parent_folder_id == folder_id
            ).order_by(Folder.name)
        )
        subfolders = subfolder_result.scalars().all()

        # Get files in this folder
        files_result = await db.execute(
            select(File).where(File.folder_id == folder_id)
        )
        files = files_result.scalars().all()

        # Build children
        children = []

        # Add subfolders recursively
        for subfolder in subfolders:
            child_node = await build_folder_node(subfolder.id)
            if child_node:
                children.append(child_node)

        # Add files
        for file in files:
            file_node = {
                "id": file.id,
                "name": file.name,
                "type": "file",
                "fileType": file.file_type,
                "path": file.path,
                "description": file.description,
                "size_bytes": file.size_bytes,
                "mime_type": file.mime_type
            }
            # Add metadata if present
            if file.tags:
                file_node["metadata"] = file.tags
            children.append(file_node)

        return {
            "id": folder.id,
            "name": folder.name,
            "type": "folder",
            "path": folder.path,
            "description": folder.description,
            "children": children
        }

    # Build root level
    tree = {
        "id": "root",
        "name": "Project Files",
        "type": "folder",
        "path": f"project_{project_id}",
        "children": []
    }

    for folder in root_folders:
        folder_node = await build_folder_node(folder.id)
        if folder_node:
            tree["children"].append(folder_node)

    # Get root-level files (no folder)
    files_result = await db.execute(
        select(File).where(
            File.project_id == project_id,
            File.folder_id.is_(None)
        )
    )
    root_files = files_result.scalars().all()

    for file in root_files:
        file_node = {
            "id": file.id,
            "name": file.name,
            "type": "file",
            "fileType": file.file_type,
            "path": file.path,
            "description": file.description,
            "size_bytes": file.size_bytes,
            "mime_type": file.mime_type
        }
        if file.tags:
            file_node["metadata"] = file.tags
        tree["children"].append(file_node)

    return tree


async def rename_folder(
    db: AsyncSession,
    folder_id: str,
    new_name: str
) -> Folder:
    """Rename a folder."""
    if not new_name or not new_name.strip():
        raise FileServiceError(status_code=400, detail="Folder name cannot be empty")

    folder = await db.get(Folder, folder_id)
    if not folder:
        raise FileServiceError(status_code=404, detail=f"Folder {folder_id} not found")

    # Check if name already exists in same parent
    parent_id = folder.parent_folder_id
    existing = await db.execute(
        select(Folder).where(
            Folder.parent_folder_id == parent_id,
            Folder.name == new_name.strip(),
            Folder.id != folder_id
        )
    )
    if existing.scalar_one_or_none():
        raise FileServiceError(
            status_code=400,
            detail=f"Folder '{new_name}' already exists in this location"
        )

    old_path = folder.path
    folder.name = new_name.strip()

    # Update path (rebuild from parent or project root)
    if parent_id:
        parent = await db.get(Folder, parent_id)
        folder.path = f"{parent.path}/{folder.name}"
    else:
        folder.path = f"project_{folder.project_id}/{folder.name}"

    await db.commit()
    await db.refresh(folder)

    logger.info(f"Renamed folder: {old_path} -> {folder.path}")
    return folder


async def delete_folder(
    db: AsyncSession,
    folder_id: str,
    recursive: bool = False
) -> None:
    """Delete a folder.

    If recursive=True, delete all contents. Otherwise, only delete if empty.
    """
    folder = await db.get(Folder, folder_id)
    if not folder:
        raise FileServiceError(status_code=404, detail=f"Folder {folder_id} not found")

    # Check if folder has contents
    subfolder_result = await db.execute(
        select(Folder).where(Folder.parent_folder_id == folder_id)
    )
    subfolder_count = len(subfolder_result.scalars().all())

    file_result = await db.execute(
        select(File).where(File.folder_id == folder_id)
    )
    file_count = len(file_result.scalars().all())

    if not recursive and (subfolder_count > 0 or file_count > 0):
        raise FileServiceError(
            status_code=400,
            detail=f"Folder is not empty ({subfolder_count} subfolders, {file_count} files). Use recursive=True to delete."
        )

    # Recursively delete contents
    if recursive:
        # Delete subfolders
        subfolders = await db.execute(
            select(Folder).where(Folder.parent_folder_id == folder_id)
        )
        for subfolder in subfolders.scalars().all():
            await delete_folder(db, subfolder.id, recursive=True)

        # Delete files
        files = await db.execute(
            select(File).where(File.folder_id == folder_id)
        )
        for file in files.scalars().all():
            await delete_file(db, file.id)

    # Delete the folder
    await db.delete(folder)
    await db.commit()

    logger.info(f"Deleted folder: {folder.path}")


async def move_file(
    db: AsyncSession,
    file_id: str,
    target_folder_id: Optional[str] = None
) -> File:
    """Move a file to a different folder (or to root)."""
    file = await db.get(File, file_id)
    if not file:
        raise FileServiceError(status_code=404, detail=f"File {file_id} not found")

    # If target folder specified, verify it exists
    if target_folder_id:
        target_folder = await db.get(Folder, target_folder_id)
        if not target_folder:
            raise FileServiceError(status_code=404, detail=f"Target folder {target_folder_id} not found")

        # Verify folder belongs to same project
        if target_folder.project_id != file.project_id:
            raise FileServiceError(
                status_code=400,
                detail="Cannot move file to a different project"
            )
    else:
        target_folder = None

    old_path = file.path
    file.folder_id = target_folder_id

    # Update file path
    if target_folder:
        file.path = f"{target_folder.path}/{file.name}"
    else:
        file.path = f"project_{file.project_id}/{file.name}"

    # Note: For cloud storage, moving files between folders would require
    # re-uploading with new storage key. For now, we only update the database path.
    # The storage_path remains the same - this is a limitation we can address later.

    await db.commit()
    await db.refresh(file)

    logger.info(f"Moved file: {old_path} -> {file.path}")
    return file


async def get_file_download_url(
    db: AsyncSession,
    file_id: str,
    expires_in: int = 3600
) -> dict:
    """
    Get download URL for a file.

    For S3/R2: Returns presigned URL
    For local: Returns None (caller should use direct file serving)

    Args:
        db: Database session
        file_id: File ID
        expires_in: URL expiration time in seconds (default 3600)

    Returns:
        Dict with 'url' and 'expires_in' keys, or None for local storage
    """
    from storage_service import get_storage

    file = await db.get(File, file_id)
    if not file:
        raise FileServiceError(status_code=404, detail=f"File {file_id} not found")

    storage = get_storage()
    backend_type = type(storage).__name__

    if backend_type == "S3StorageBackend":
        # Extract storage key from storage_path
        storage_key = file.storage_path.replace("uploads/", "") if file.storage_path.startswith("uploads/") else file.storage_path

        # Generate presigned URL
        url = await storage.get_file_url(storage_key, expires_in=expires_in)
        return {
            "url": url,
            "expires_in": expires_in,
            "filename": file.name
        }
    else:
        # Local storage - return None to indicate direct file serving
        return None
